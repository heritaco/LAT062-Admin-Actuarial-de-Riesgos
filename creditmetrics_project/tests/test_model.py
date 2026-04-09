from openpyxl import load_workbook

from creditmetrics_project.model import run_analysis


def test_convolution_matches_bruteforce() -> None:
    result = run_analysis(nr_policy="renormalize")
    portfolio = result["portfolio_summary"]
    assert portfolio.convolution_matches_bruteforce_cents is True
    assert portfolio.support_size_bruteforce == 18 ** 3
    assert portfolio.support_size_cents <= portfolio.support_size_bruteforce


def test_var_is_stable() -> None:
    result = run_analysis(nr_policy="renormalize")
    portfolio = result["portfolio_summary"]
    assert round(portfolio.var_999, 2) == 73.81


def test_outputs_include_excel_and_bruteforce_tables(tmp_path) -> None:
    run_analysis(nr_policy="renormalize", output_dir=tmp_path)

    xlsx_path = tmp_path / "resultados_creditmetrics.xlsx"
    pdf_path = tmp_path / "reporte_creditmetrics.pdf"
    figures_path = tmp_path / "figuras"

    assert xlsx_path.exists()
    assert pdf_path.exists()
    assert pdf_path.stat().st_size > 0
    assert figures_path.exists()
    assert (figures_path / "01_tabla_resumen_portafolio.png").exists()
    assert (figures_path / "06_grafica_portafolio_distribucion.png").exists()
    assert (figures_path / "10_tabla_comparacion_convolucion_vs_bruteforce.png").exists()

    workbook = load_workbook(xlsx_path, data_only=True)
    assert "resumen_portafolio" in workbook.sheetnames
    assert "portafolio_distrib" in workbook.sheetnames
    assert "bruteforce_pmf" in workbook.sheetnames
    assert "comparacion_pmf" in workbook.sheetnames
    assert "bruteforce_escenarios" in workbook.sheetnames
    assert "convolucion_2_bonos" in workbook.sheetnames
    assert "convolucion_3_bonos" in workbook.sheetnames

    comparison_sheet = workbook["comparacion_pmf"]
    headers = [cell.value for cell in comparison_sheet[1]]
    assert headers == [
        "portfolio_value",
        "probability_convolution",
        "probability_bruteforce",
        "difference",
        "cumulative_convolution",
        "cumulative_bruteforce",
    ]

    brute_sheet = workbook["bruteforce_escenarios"]
    assert brute_sheet.max_row == 18 ** 3 + 1
